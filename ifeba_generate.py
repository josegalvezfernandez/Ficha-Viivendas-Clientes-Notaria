import openpyxl
from pathlib import Path

# Paths (edit if needed)
SOURCE_PATH = Path(r"C:\Users\jg123\OneDrive\Desktop\CUADRO MAESTRO IFEBA 1 IA.xlsx")
DESKTOP_DIR = Path(r"C:\Users\jg123\OneDrive\Desktop")
OUTPUT_PATH = Path(r"C:\Users\jg123\OneDrive\Desktop\Ifeba 1.xlsx")

START_ROW = 12
MAX_ROWS = None  # all viviendas
TEMPLATE_SHEET = "Hoja1"


def num(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def main():
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Source not found: {SOURCE_PATH}")
    template_path = None
    if DESKTOP_DIR.exists():
        for f in DESKTOP_DIR.iterdir():
            if f.suffix.lower() == ".xlsx" and "Ficha Vvda" in f.name:
                template_path = f
                break
    if not template_path:
        raise FileNotFoundError("Template not found on Desktop (Ficha Vvda*.xlsx)")

    # Read source workbook
    src_wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    src_ws = src_wb["SITUACION COMERCIAL"] if "SITUACION COMERCIAL" in src_wb.sheetnames else src_wb.active
    mejoras_ws = src_wb["MEJORAS"] if "MEJORAS" in src_wb.sheetnames else None

    # Load template
    tmpl_wb = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in tmpl_wb.sheetnames:
        raise ValueError(f"Template sheet not found: {TEMPLATE_SHEET}")

    base_sheet = tmpl_wb[TEMPLATE_SHEET]
    created = 0

    last_row = src_ws.max_row if MAX_ROWS is None else (START_ROW + MAX_ROWS - 1)
    for row in range(START_ROW, last_row + 1):
        v = {}
        v["orden"] = src_ws["E" + str(row)].value
        if v["orden"] is None or str(v["orden"]).strip() == "":
            continue

        v["portal"] = src_ws["F" + str(row)].value
        v["planta"] = src_ws["G" + str(row)].value
        v["letra"] = src_ws["H" + str(row)].value
        v["garaje"] = src_ws["J" + str(row)].value
        v["trastero"] = src_ws["K" + str(row)].value
        v["titular1"] = src_ws["AB" + str(row)].value
        v["titular2"] = src_ws["AF" + str(row)].value
        v["situacion"] = src_ws["AJ" + str(row)].value
        v["dni1"] = src_ws["AC" + str(row)].value
        v["dni2"] = src_ws["AG" + str(row)].value
        v["direccion"] = src_ws["AK" + str(row)].value
        v["sup_construida"] = src_ws["Q" + str(row)].value
        v["sup_util"] = src_ws["P" + str(row)].value
        v["neto"] = src_ws["S" + str(row)].value
        v["iva"] = src_ws["T" + str(row)].value
        v["total"] = src_ws["U" + str(row)].value
        v["reserva"] = src_ws["V" + str(row)].value
        v["compraventa"] = src_ws["W" + str(row)].value
        v["durante_obra"] = src_ws["X" + str(row)].value
        v["a_llaves"] = src_ws["Y" + str(row)].value
        v["hipoteca"] = src_ws["Z" + str(row)].value
        v["finca_registral"] = src_ws["AA" + str(row)].value

        if mejoras_ws:
            # Match by Nº Orden in column B of MEJORAS sheet (row 9 is first vivienda)
            mej_row = None
            for r in range(9, mejoras_ws.max_row + 1):
                if mejoras_ws["B" + str(r)].value == v["orden"]:
                    mej_row = r
                    break
            if mej_row:
                v["mej_f15"] = mejoras_ws["Q" + str(mej_row)].value
                v["mej_g15"] = num(mejoras_ws["O" + str(mej_row)].value) + num(mejoras_ws["H" + str(mej_row)].value)
                v["mej_h15"] = mejoras_ws["R" + str(mej_row)].value
                v["mejoras_h20"] = mejoras_ws["R" + str(mej_row)].value
            else:
                v["mej_f15"] = None
                v["mej_g15"] = None
                v["mej_h15"] = None
                v["mejoras_h20"] = None
        else:
            v["mej_f15"] = None
            v["mej_g15"] = None
            v["mej_h15"] = None
            v["mejoras_h20"] = None

        new_sheet = tmpl_wb.copy_worksheet(base_sheet)
        sheet_name = str(v["orden"]).strip()
        new_sheet.title = sheet_name
        new_sheet.sheet_state = "visible"

        # Fill template cells
        new_sheet["E6"].value = v["orden"]
        new_sheet["F6"].value = v["portal"]
        new_sheet["G6"].value = v["planta"]
        new_sheet["H6"].value = v["letra"]
        new_sheet["I6"].value = v["garaje"]
        new_sheet["J6"].value = v["trastero"]

        titular1 = v["titular1"] or ""
        titular2 = v["titular2"] or ""
        titulares = f"{titular1} / {titular2}".strip(" /")
        new_sheet["C9"].value = titulares
        new_sheet["C10"].value = v["situacion"]
        new_sheet["C11"].value = v["dni1"]
        new_sheet["C12"].value = v["dni2"]
        new_sheet["C13"].value = v["direccion"]

        new_sheet["F9"].value = v["sup_construida"]
        new_sheet["H9"].value = v["sup_util"]

        new_sheet["F14"].value = v["neto"]
        new_sheet["G14"].value = v["iva"]
        new_sheet["H14"].value = v["total"]
        new_sheet["F15"].value = v["mej_f15"]
        new_sheet["G15"].value = v["mej_g15"]
        new_sheet["H15"].value = v["mej_h15"]

        # Totals row
        new_sheet["F16"].value = num(v["mej_f15"]) + num(v["neto"])
        new_sheet["G16"].value = num(v["mej_g15"]) + num(v["iva"])
        new_sheet["H16"].value = num(v["mej_h15"]) + num(v["total"])

        h19 = num(v["reserva"]) + num(v["compraventa"]) + num(v["durante_obra"])
        new_sheet["H18"].value = None
        new_sheet["H19"].value = None if abs(h19) < 0.0001 else h19
        new_sheet["H20"].value = v["mejoras_h20"]
        new_sheet["H21"].value = v["a_llaves"]
        new_sheet["H22"].value = v["hipoteca"]
        new_sheet["H24"].value = "=SUM(H19:H22)"
        new_sheet["F2"].value = v["finca_registral"]

        # Totals mismatch warnings disabled per request

        created += 1

    src_wb.close()

    # Remove original template sheets only if we created at least one sheet
    if created > 0:
        for name in list(tmpl_wb.sheetnames):
            if name in ("Hoja1", "Hoja2", "Hoja3"):
                del tmpl_wb[name]

        # Ensure at least one sheet is visible
        if not any(ws.sheet_state == "visible" for ws in tmpl_wb.worksheets):
            tmpl_wb.worksheets[0].sheet_state = "visible"
    else:
        # Fallback: keep template sheets if no data rows were created
        for ws in tmpl_wb.worksheets:
            ws.sheet_state = "visible"

    tmpl_wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
